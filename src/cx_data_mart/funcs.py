"""Helper functions."""

import pathlib
import re
import unicodedata
from typing import Any, Literal

import chardet
import numpy as np
import pandas as pd
from janitor import (
    clean_names,
    drop_constant_columns,
    drop_duplicate_columns,
    remove_empty,
)
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from pandas._libs.missing import NAType
from pydantic import ConfigDict, validate_call


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def remove_whitespaces_str_col(
    df: pd.DataFrame,
    text_cols: list[str],
    action: Literal[
        "leading_and_trailing",
        "remove_all_whitespace",
        "single_space",
    ] = "single_space",
) -> pd.DataFrame:
    """Remove or standardize whitespace in specified string columns of a DataFrame."""
    for col in text_cols:
        assert col in df.columns, f"Column {col} not in DataFrame"
        if action == "leading_and_trailing":
            df[col] = df[col].str.strip()
        elif action == "remove_all_whitespace":
            df[col] = df[col].str.replace(r"\s+", "", regex=True)
        elif action == "single_space":
            df[col] = df[col].str.split().str.join(" ")
    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def lowercase_str_col(df: pd.DataFrame, text_cols: list[str]) -> pd.DataFrame:
    """Converts text to lowercase in a pandas dataframe.

    Args:
        df: The input Pandas DataFrame.
        text_cols: List of column names whose string values should be lowercased.


    Returns:
        The DataFrame with specified columns lowercased.

    """
    for col in text_cols:
        if col not in df.columns:
            print(f"Warning: Column '{col}' not found in DataFrame. Skipping.")
            continue
        df[col] = df[col].astype("string").str.lower()
    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def remove_accents_and_special_chars(
    df: pd.DataFrame,
    text_cols: list[str],
) -> pd.DataFrame:
    """Removes accents and common special characters from text.

    Args:
        df: The input Pandas DataFrame.
        text_cols: List of column names whose string values should be processed.

    Returns:
        The DataFrame with specified columns cleaned
    """

    def clean_string(text: str) -> str | NAType:
        """Helper function to clean a single string.

        1. Handles NaN values gracefully.
        2. Converts input to string.
        3. Removes accents.
        4. Removes non-alphanumeric/non-whitespace characters.
        5. Normalizes internal whitespace and strips leading/trailing.
        """
        # Handle pandas NA values (from 'string' dtype)
        if pd.isna(text):
            return pd.NA

        # Ensure text is a string (e.g., if it was a number/boolean converted to string)
        text = str(text)

        # 1. Remove accents (diacritics)
        # NFKD: Normalization Form Compatibility Decomposition. Decomposes characters
        # into their base character and diacritics.
        # encode('ascii', 'ignore'): Encodes to ASCII, ignoring characters that can't be
        # represented (i.e., the diacritics).
        # decode('utf-8'): Decodes back to a UTF-8 string.
        text = (
            unicodedata.normalize("NFKD", text)
            .encode("ascii", "ignore")
            .decode("utf-8")
        )

        # 2. Remove special characters (keep alphanumeric and whitespace)
        # [^a-zA-Z0-9\s]: Matches any character that is NOT a letter, NOT a number,
        # NOT a whitespace character.
        text = re.sub(r"[^a-zA-Z0-9\s]", "", text)

        # 3. Normalize whitespace (collapse multiple spaces/tabs/newlines into single
        # space, then strip)
        text = re.sub(r"\s+", " ", text).strip()

        return text

    for col in text_cols:
        if col not in df.columns:
            print(
                f"Warning: Column '{col}' not found in DataFrame. Skipping.",
            )
            continue

        # Convert column to 'string' dtype first to ensure .str accessor works
        # and handles non-string data by converting them to their string representation.
        # This also ensures pd.NA for missing values.
        df[col] = df[col].astype("string").apply(clean_string)

    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def set_pandas_setup() -> None:
    """Pandas setup."""
    pd.options.display.width = None
    pd.options.display.max_columns = None
    pd.set_option("display.max_rows", 3000)
    pd.set_option("display.max_columns", 3000)


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def convert_column_to_integers(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Convert a pandas column to integers.

    Args:
        df: Dataframe.
        column: Column to convert to integers.

    Returns:
        Dataframe with column converted to integers.

    References:
        https://stackoverflow.com/questions/62899860/how-can-i-resolve-typeerror-cannot-safely-cast-non-equivalent-float64-to-int6

    """
    df[column] = df[column].astype("Int64")
    df[column] = np.floor(pd.to_numeric(df[column], errors="coerce")).astype("Int64")
    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def add_type_to_pd_cols(dtype: dict[str, str], df: pd.DataFrame) -> pd.DataFrame:
    """Add data types to columns in a pandas dataframe."""
    for column, dtype_str in dtype.items():
        if dtype_str == "Int64":
            df = convert_column_to_integers(df=df, column=column)
        else:
            df[column] = df[column].astype(dtype_str)

    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def read_pd_dataset(  # noqa: PLR0915, PLR0917
    fpath: str | pathlib.Path,
    dtype: dict[str, Any] | None = None,
    true_values: list[str] | None = None,
    false_values: list[str] | None = None,
    date_columns: list[str] | None = None,
    na_values: str | tuple[str] = (" ", "  ", "-", "?"),
    *,
    drop_empty_columns: bool = True,
) -> pd.DataFrame:
    """Read dataset file with protections."""
    fpath = str(fpath)
    if fpath.endswith(("xls", "xlsx")):
        if is_excel_file_open(fpath=fpath):
            raise ValueError(f"Excel file {fpath} is open, please close it.")
        if excel_has_multiple_sheets(fpath=fpath):
            raise ValueError(
                "Excel file has multiple sheets. Please work with one sheet per file.",
            )
        if has_columns_with_same_name(fpath=fpath):
            raise ValueError("Excel file has columns with the same name.")
        if excel_file_has_formulas(fpath=fpath):
            raise ValueError("Excel file has formulas, remove them from file.")
        unmerge_cells_in_excel_file(input_file=fpath, output_file=fpath)
        try:
            df = pd.read_excel(
                fpath,
                dtype=dtype,
                true_values=true_values,
                false_values=false_values,
                na_values=na_values,
                parse_dates=date_columns,
            )
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Error: '{fpath}' not found.") from e
        except OSError as e:
            raise RuntimeError(f"Error: Failed to read '{fpath}'.") from e
    elif fpath.endswith(".csv"):
        encoding = get_text_encoding(fpath=fpath)
        if not encoding.startswith("UTF-8"):
            raise ValueError(
                f"Please convert CSV file ({fpath}) to UTF-8, got {encoding}",
            )
        if has_columns_with_same_name(fpath=fpath):
            raise ValueError(f"CSV file ({fpath}) has columns with the same name.")
        try:
            df = pd.read_csv(
                fpath,
                dtype=dtype,
                engine="pyarrow",
                on_bad_lines="warn",
                true_values=true_values,
                false_values=false_values,
                parse_dates=date_columns,
                na_values=na_values,
            )
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Error: '{fpath}' not found.") from e
        except OSError as e:
            raise RuntimeError(f"Error: Failed to read '{fpath}'.") from e
    elif fpath.endswith(".feather"):
        try:
            df = pd.read_feather(path=fpath)
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Error: '{fpath}' not found.") from e
        except OSError as e:
            raise RuntimeError(f"Error: Failed to read '{fpath}'.") from e
    elif fpath.endswith(".parquet"):
        try:
            df = pd.read_parquet(path=fpath, engine="pyarrow")
        except FileNotFoundError as e:
            raise FileNotFoundError(f"Error: '{fpath}' not found.") from e
        except OSError as e:
            raise RuntimeError(f"Error: Failed to read '{fpath}'.") from e
    else:
        raise OSError("could not read file, check extension.")

    # Drop all columns and rows that are completely empty.
    if drop_empty_columns:
        init_cols = df.shape[1]
        df = remove_empty(df=df)
        end_cols = df.shape[1]
        if init_cols != end_cols:
            print(f"Empty columns dropped: {init_cols - end_cols}!")

    # Check mixed types in one column
    object_col_types = df.select_dtypes("object")
    if object_col_types.shape[1] != 0:
        columns_with_mixed_types = df.pipe(get_mixed_columns)
        if (
            columns_with_mixed_types is not None
            and columns_with_mixed_types.shape[0] != 0
        ):
            print(
                f"Mixed types in columns: {columns_with_mixed_types}!!!",
            )

    # Drop duplicates
    df = df.drop_duplicates()
    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def get_mixed_columns(df: pd.DataFrame) -> list[str] | None:
    """Find columns in dataframe with mixed types.

    Args:
        df: dataframe

    References:
        https://stackoverflow.com/questions/72550219/how-to-deal-with-a-dataframe-containing-columns-with-mixed-types

    """
    mixed_columns = (
        df.select_dtypes("object")
        .apply(pd.api.types.infer_dtype)
        .loc[lambda x: x.str.contains("mixed")]
    )
    return mixed_columns


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def is_excel_file_open(fpath: str) -> bool:
    """Check if an Excel file is open."""
    try:
        # Try to open the file in exclusive mode
        with open(fpath, "r+", encoding="utf-8"):
            return False
    except OSError:
        return True


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def excel_has_multiple_sheets(fpath: str) -> bool:
    """Check if Excel file has more than one sheet."""
    # Load the Excel file
    xls = pd.ExcelFile(fpath)

    # Get the number of sheets
    sheet_names = xls.sheet_names
    num_sheets = len(sheet_names)

    # Check if there are more than one sheet
    max_number_of_sheets = 1
    return num_sheets > max_number_of_sheets


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def has_columns_with_same_name(fpath: str) -> bool:
    """Check if a file has columns with the same name."""
    if fpath.endswith(("xls", "xlsx")):
        cols_per_name = (
            pd.read_excel(fpath, header=None, nrows=1).iloc[0, :].value_counts()
        )
    elif fpath.endswith(".csv"):
        cols_per_name = (
            pd.read_csv(fpath, header=None, nrows=1).iloc[0, :].value_counts()
        )
    else:
        raise ValueError("File format not supported.")
    duplicated_header_names = cols_per_name[cols_per_name > 1].to_dict()
    if duplicated_header_names:
        print(
            f"Warning: columns: {duplicated_header_names} "
            f"with the same name found in file: {fpath}",
        )

        return True
    return False


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def excel_file_has_formulas(fpath: str) -> bool:
    """Check if an Excel file has formulas."""
    wb = load_workbook(fpath, data_only=False)  # Set data_only=False to access formulas

    # Iterate through each sheet in the workbook
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":  # Check if the cell contains a formula
                    return True
    return False


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def unmerge_cells_in_excel_file(input_file: str, output_file: str) -> None:
    """Unmerge cells in an Excel file.

    Args:
        input_file: input file path.
        output_file: output file path.

    References:
        https://stackoverflow.com/questions/47872349/how-to-split-merged-excel-cells-with-python

    """
    wb = load_workbook(filename=input_file)

    for st_name in wb.sheetnames:
        st = wb[st_name]
        mcr_coord_list = [mcr.coord for mcr in st.merged_cells.ranges]

        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = st.cell(row=min_row, column=min_col).value
            st.unmerge_cells(mcr)
            for row in st.iter_rows(
                min_col=min_col,
                min_row=min_row,
                max_col=max_col,
                max_row=max_row,
            ):
                for cell in row:
                    cell.value = top_left_cell_value

    number_of_rows_start = pd.read_excel(input_file).shape[0]
    wb.save(output_file)
    number_of_rows_end = pd.read_excel(input_file).shape[0]
    assert number_of_rows_start == number_of_rows_end, (
        "Number of rows changed after unmerging cells."
    )


@validate_call
def get_text_encoding(fpath: str) -> str:
    """Get text encoding."""
    return chardet.detect(pathlib.Path(fpath).read_bytes())["encoding"]


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def janitor_df_cleaning(
    df: pd.DataFrame,
    truncate_limit: int | None = None,
    *,
    drop_empty_cols: bool = False,
    drop_duplicated_cols: bool = False,
) -> pd.DataFrame:
    """Clean dataframe using janitor library functions.

    Args:
        df: Input pandas DataFrame to clean.
        truncate_limit: Maximum length for column names after cleaning.
        drop_empty_cols: Whether to drop constant and empty columns.
        drop_duplicated_cols: Whether to drop duplicate columns.

    Returns:
        Cleaned DataFrame with standardized column names and
            optionally removed empty/duplicate columns.
    """
    assert df.shape[0] > 0, "Empty dataframe cannot be cleaned."
    df = clean_names(df=df, truncate_limit=truncate_limit, remove_special=True)
    df.columns = df.columns.str.rstrip("_")  # Remove trailing underscores
    if drop_empty_cols:
        df = drop_constant_columns(df=df)
        df = remove_empty(df=df)
    if drop_duplicated_cols and has_duplicate_columns(df=df):
        duplicated_columns = get_duplicated_columns(df=df)
        for duplicated_column in duplicated_columns:
            df = drop_duplicate_columns(df=df, column_name=duplicated_column)
    return df


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def has_duplicate_columns(df: pd.DataFrame) -> bool:
    """Check if a DataFrame has duplicate columns."""
    duplicated_columns = get_duplicated_columns(df=df)
    if duplicated_columns:
        print("Warning: duplicate columns found:", duplicated_columns)
        return True
    return False


@validate_call(config=ConfigDict(arbitrary_types_allowed=True))
def get_duplicated_columns(df: pd.DataFrame) -> list[str] | None:
    """Get duplicated columns in a DataFrame."""
    duplicated_columns = df.columns[df.columns.duplicated()].unique().tolist()
    if len(duplicated_columns) > 0:
        return duplicated_columns
    return None
